"""国家模块视图。"""

from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.utils import timezone
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    OpenApiTypes,
    extend_schema,
    extend_schema_view,
    inline_serializer,
)
from openpyxl import load_workbook
from rest_framework import filters, serializers, status
from rest_framework.generics import ListAPIView, RetrieveAPIView
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet
from rest_framework.views import APIView

from apps.countries.models import Country, CountryIndicator
from apps.countries.serializers import (
    CountryDetailSerializer,
    CountryIndicatorSerializer,
    CountrySerializer,
)
from apps.recommendation.recommendation import TravelRecommendationEngine
from apps.recommendation.tourism_suitability import TourismSuitabilityCalculator
from apps.system.models import OperationLog, record_operation_log
from common.permissions import IsAdminRole


IMPORT_REQUIRED_HEADERS = {
    "国家名称": "country_name",
    "安全指数": "safety_index",
    "幸福指数": "happiness_index",
    "消费指数": "cost_index",
    "旅游适宜指数": "tourism_index",
    "洲别": "continent",
}

IMPORT_SCORE_FIELDS = {
    "safety_index": "安全指数",
    "happiness_index": "幸福指数",
    "cost_index": "消费指数",
    "tourism_index": "旅游适宜指数",
}


def normalize_excel_cell(value):
    """将 Excel 单元格值标准化为便于校验的字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def parse_score(value, field_name, row_number, errors):
    """解析并校验 0 到 100 的指数值。"""
    raw_value = normalize_excel_cell(value)
    if not raw_value:
        errors.append(
            {
                "row": row_number,
                "field": field_name,
                "message": f"{field_name}不能为空",
            }
        )
        return None

    try:
        score = Decimal(raw_value)
    except (InvalidOperation, ValueError):
        errors.append(
            {
                "row": row_number,
                "field": field_name,
                "message": f"{field_name}必须是数字",
            }
        )
        return None

    if score < Decimal("0") or score > Decimal("100"):
        errors.append(
            {
                "row": row_number,
                "field": field_name,
                "message": f"{field_name}必须在 0 到 100 之间",
            }
        )
        return None

    return score.quantize(Decimal("0.01"))


def build_import_country_code(country_name):
    """为 Excel 新增国家生成稳定且不易冲突的国家代码。"""
    import hashlib

    digest = hashlib.sha1(country_name.encode("utf-8")).hexdigest().upper()
    base_code = f"IMP{digest[:7]}"
    if not Country.objects.filter(code=base_code).exists():
        return base_code

    for index in range(1, 100):
        candidate = f"IMP{digest[:5]}{index:02d}"
        if not Country.objects.filter(code=candidate).exists():
            return candidate

    return f"IMP{digest[:4]}ZZ"


def read_indicator_import_rows(excel_file):
    """读取并校验国家指标 Excel，返回待写入的数据与错误列表。"""
    errors = []
    try:
        workbook = load_workbook(excel_file, read_only=True, data_only=True)
    except Exception:
        return [], [
            {
                "row": 0,
                "field": "file",
                "message": "Excel 文件读取失败，请上传有效的 .xlsx 文件",
            }
        ]

    worksheet = workbook.active
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [normalize_excel_cell(cell) for cell in (header_row or [])]
    header_index_map = {header: index for index, header in enumerate(headers) if header}
    missing_headers = [
        header for header in IMPORT_REQUIRED_HEADERS if header not in header_index_map
    ]
    if missing_headers:
        return [], [
            {
                "row": 1,
                "field": "表头",
                "message": f"缺少必填列：{'、'.join(missing_headers)}",
            }
        ]

    rows = []
    seen_country_names = {}
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(normalize_excel_cell(cell) for cell in row):
            continue

        item = {}
        country_name = normalize_excel_cell(row[header_index_map["国家名称"]])
        continent = normalize_excel_cell(row[header_index_map["洲别"]])

        if not country_name:
            errors.append(
                {
                    "row": row_number,
                    "field": "国家名称",
                    "message": "国家名称不能为空",
                }
            )
        else:
            normalized_country_name = country_name.casefold()
            if normalized_country_name in seen_country_names:
                errors.append(
                    {
                        "row": row_number,
                        "field": "国家名称",
                        "message": f"国家名称与第 {seen_country_names[normalized_country_name]} 行重复",
                    }
                )
            else:
                seen_country_names[normalized_country_name] = row_number

        if not continent:
            errors.append(
                {
                    "row": row_number,
                    "field": "洲别",
                    "message": "洲别不能为空",
                }
            )

        item["country_name"] = country_name
        item["continent"] = continent

        for field_key, field_name in IMPORT_SCORE_FIELDS.items():
            score = parse_score(
                row[header_index_map[field_name]],
                field_name,
                row_number,
                errors,
            )
            item[field_key] = score

        rows.append(item)

    if not rows and not errors:
        errors.append(
            {
                "row": 0,
                "field": "file",
                "message": "Excel 中没有可导入的数据行",
            }
        )

    return rows, errors


def build_indicator_payload(indicator):
    """将国家指标记录转换为前端分析页面需要的数据结构。"""
    raw_visa_score = (
        float(indicator.visa_index)
        if float(indicator.visa_index) != 50.0
        else float(indicator.tourism_index)
    )
    country_payload = {
        "country_id": indicator.country.id,
        "country_name": indicator.country.name_zh,
        "country_name_en": indicator.country.name_en,
        "continent": indicator.country.continent,
        "raw_visa_score": raw_visa_score,
        "visa_index": float(indicator.visa_index),
        "tourism_index": float(indicator.tourism_index),
        "safety_index": float(indicator.safety_index),
        "cost_index": float(indicator.cost_index),
        "happiness_index": float(indicator.overall_score),
    }
    tourism_detail = TourismSuitabilityCalculator.build_detail(country_payload)
    country_payload["tourism_index"] = tourism_detail["tourism_index"]
    recommendation_index = TravelRecommendationEngine.calculate_default_recommendation_index(
        country_payload
    )
    return {
        "country_id": indicator.country.id,
        "country_name": indicator.country.name_zh,
        "country_name_en": indicator.country.name_en,
        "code": indicator.country.code,
        "continent": indicator.country.continent,
        "capital": indicator.country.capital,
        "language": indicator.country.language,
        "currency": indicator.country.currency,
        "year": indicator.year,
        "recommendation_index": recommendation_index,
        "tourism_index": round(float(tourism_detail["tourism_index"]), 2),
        "tourism_detail": tourism_detail,
        "safety_index": round(float(indicator.safety_index), 2),
        "ppp_index": round(float(indicator.cost_index), 2),
        "cost_index": round(float(indicator.cost_index), 2),
        "happiness_index": round(float(indicator.overall_score), 2),
        "visa_index": round(float(raw_visa_score), 2),
        "overall_score": round(float(indicator.overall_score), 2),
        "data_source": indicator.data_source,
    }


@extend_schema(
    tags=["国家数据"],
    summary="查询国家列表",
    description="管理员查询启用状态的国家基础数据，支持按洲别筛选、搜索和排序。",
    parameters=[
        OpenApiParameter("continent", str, OpenApiParameter.QUERY, description="洲别，如 亚洲、欧洲"),
        OpenApiParameter("search", str, OpenApiParameter.QUERY, description="按国家中文名、英文名、代码或洲别搜索"),
        OpenApiParameter("ordering", str, OpenApiParameter.QUERY, description="排序字段，如 name_zh 或 -created_at"),
    ],
    responses={200: CountrySerializer(many=True)},
)
class CountryListAPIView(ListAPIView):
    """国家列表查询接口。"""

    serializer_class = CountrySerializer
    permission_classes = [IsAdminRole]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["name_zh", "name_en", "code", "continent"]
    ordering_fields = ["created_at", "name_zh"]
    ordering = ["name_zh"]

    @extend_schema(
        tags=["国家数据"],
        summary="查询国家列表",
        description="管理员查询启用状态的国家基础数据，支持按洲别筛选、搜索和排序。",
        parameters=[
            OpenApiParameter("continent", str, OpenApiParameter.QUERY, description="洲别，如 亚洲、欧洲"),
            OpenApiParameter("search", str, OpenApiParameter.QUERY, description="按国家中文名、英文名、代码或洲别搜索"),
            OpenApiParameter("ordering", str, OpenApiParameter.QUERY, description="排序字段，如 name_zh 或 -created_at"),
        ],
        responses={200: CountrySerializer(many=True)},
    )
    def get_queryset(self):
        """获取国家列表查询集，并支持按洲别筛选。"""
        queryset = Country.objects.filter(is_active=True)
        continent = self.request.query_params.get("continent")
        if continent:
            queryset = queryset.filter(continent=continent)
        return queryset


@extend_schema(
    tags=["国家数据"],
    summary="查询国家详情",
    description="管理员查看单个国家基础信息及关联年度指标。",
    responses={200: CountryDetailSerializer},
)
class CountryDetailAPIView(RetrieveAPIView):
    """国家详情接口。"""

    serializer_class = CountryDetailSerializer
    permission_classes = [IsAdminRole]
    queryset = Country.objects.filter(is_active=True)


@extend_schema(
    tags=["国家数据"],
    summary="查询国家指标数据",
    description="管理员查询国家年度安全指数、幸福指数、消费指数、旅游适宜指数等指标。",
    parameters=[
        OpenApiParameter("country", int, OpenApiParameter.QUERY, description="国家 ID"),
        OpenApiParameter("year", int, OpenApiParameter.QUERY, description="指标年份"),
        OpenApiParameter("ordering", str, OpenApiParameter.QUERY, description="排序字段，如 -year、-safety_index"),
    ],
    responses={200: CountryIndicatorSerializer(many=True)},
)
class CountryIndicatorListAPIView(ListAPIView):
    """国家指标数据查询接口。"""

    serializer_class = CountryIndicatorSerializer
    permission_classes = [IsAdminRole]
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ["year", "overall_score", "safety_index", "tourism_index"]
    ordering = ["-year", "-overall_score"]

    @extend_schema(
        tags=["国家数据"],
        summary="查询国家指标数据",
        description="管理员查询国家年度安全指数、幸福指数、消费指数、旅游适宜指数等指标。",
        parameters=[
            OpenApiParameter("country", int, OpenApiParameter.QUERY, description="国家 ID"),
            OpenApiParameter("year", int, OpenApiParameter.QUERY, description="指标年份"),
            OpenApiParameter("ordering", str, OpenApiParameter.QUERY, description="排序字段，如 -year、-safety_index"),
        ],
        responses={200: CountryIndicatorSerializer(many=True)},
    )
    def get_queryset(self):
        """获取国家指标查询集，并支持按国家与年份筛选。"""
        queryset = CountryIndicator.objects.select_related("country").filter(
            country__is_active=True
        )

        country_id = self.request.query_params.get("country")
        if country_id:
            queryset = queryset.filter(country_id=country_id)

        year = self.request.query_params.get("year")
        if year:
            queryset = queryset.filter(year=year)

        return queryset


class CountryIndicatorImportExcelAPIView(APIView):
    """国家指标 Excel 批量导入接口。"""

    permission_classes = [IsAdminRole]
    parser_classes = [MultiPartParser, FormParser]

    @extend_schema(
        tags=["国家数据"],
        summary="导入国家指标 Excel",
        description=(
            "管理员上传 .xlsx 文件批量导入国家指标数据。Excel 表头必须包含："
            "国家名称、安全指数、幸福指数、消费指数、旅游适宜指数、洲别。"
            "导入前会校验空值、重复国家和 0-100 数值范围。"
        ),
        request={
            "multipart/form-data": inline_serializer(
                name="CountryIndicatorImportExcelRequest",
                fields={
                    "file": serializers.FileField(help_text="仅支持 .xlsx 文件"),
                    "year": serializers.IntegerField(required=False, help_text="指标年份，默认当前年份"),
                },
            )
        },
        responses={
            200: OpenApiResponse(description="导入成功，返回新增/更新统计"),
            400: OpenApiResponse(description="文件格式、年份或 Excel 内容校验失败"),
        },
        examples=[
            OpenApiExample(
                "Excel 表头示例",
                value={
                    "headers": ["国家名称", "安全指数", "幸福指数", "消费指数", "旅游适宜指数", "洲别"],
                    "row": ["日本", 88, 74, 62, 91, "亚洲"],
                },
                response_only=True,
            )
        ],
    )
    def post(self, request):
        """上传 Excel 文件并批量创建或更新国家指标。"""
        excel_file = request.FILES.get("file")
        if not excel_file:
            record_operation_log(
                request,
                OperationLog.OPERATION_EXCEL_IMPORT,
                operation_object="未上传文件",
                operation_result=OperationLog.RESULT_FAILED,
                detail="导入失败：请上传 Excel 文件",
            )
            return Response(
                {
                    "code": 400,
                    "message": "请上传 Excel 文件",
                    "data": None,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not excel_file.name.lower().endswith(".xlsx"):
            record_operation_log(
                request,
                OperationLog.OPERATION_EXCEL_IMPORT,
                operation_object=excel_file.name,
                operation_result=OperationLog.RESULT_FAILED,
                detail="导入失败：当前仅支持 .xlsx 格式的 Excel 文件",
            )
            return Response(
                {
                    "code": 400,
                    "message": "当前仅支持 .xlsx 格式的 Excel 文件",
                    "data": None,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        year = request.data.get("year") or timezone.localdate().year
        try:
            year = int(year)
        except (TypeError, ValueError):
            record_operation_log(
                request,
                OperationLog.OPERATION_EXCEL_IMPORT,
                operation_object=excel_file.name,
                operation_result=OperationLog.RESULT_FAILED,
                detail="导入失败：年份必须是数字",
            )
            return Response(
                {
                    "code": 400,
                    "message": "年份必须是数字",
                    "data": None,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if year < 2000 or year > 2100:
            record_operation_log(
                request,
                OperationLog.OPERATION_EXCEL_IMPORT,
                operation_object=excel_file.name,
                operation_result=OperationLog.RESULT_FAILED,
                detail="导入失败：年份必须在 2000 到 2100 之间",
            )
            return Response(
                {
                    "code": 400,
                    "message": "年份必须在 2000 到 2100 之间",
                    "data": None,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        rows, errors = read_indicator_import_rows(excel_file)
        if errors:
            record_operation_log(
                request,
                OperationLog.OPERATION_EXCEL_IMPORT,
                operation_object=excel_file.name,
                operation_result=OperationLog.RESULT_FAILED,
                detail=f"导入校验失败：共 {len(errors)} 个错误",
            )
            return Response(
                {
                    "code": 400,
                    "message": "Excel 校验失败，请修正后重新上传",
                    "data": {
                        "year": year,
                        "total_rows": len(rows),
                        "required_headers": list(IMPORT_REQUIRED_HEADERS.keys()),
                        "errors": errors,
                    },
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created_countries = 0
        updated_countries = 0
        created_indicators = 0
        updated_indicators = 0
        import_source = f"Excel 批量导入（{timezone.localdate().isoformat()}）"

        with transaction.atomic():
            for item in rows:
                country = (
                    Country.objects.filter(
                        Q(name_zh=item["country_name"]) | Q(name_en=item["country_name"])
                    )
                    .order_by("id")
                    .first()
                )

                if country is None:
                    country = Country.objects.create(
                        name_zh=item["country_name"],
                        name_en=item["country_name"],
                        code=build_import_country_code(item["country_name"]),
                        continent=item["continent"],
                        is_active=True,
                    )
                    created_countries += 1
                else:
                    changed = False
                    if country.continent != item["continent"]:
                        country.continent = item["continent"]
                        changed = True
                    if not country.is_active:
                        country.is_active = True
                        changed = True
                    if changed:
                        country.save(update_fields=["continent", "is_active", "updated_at"])
                        updated_countries += 1

                indicator = CountryIndicator.objects.filter(
                    country=country,
                    year=year,
                ).first()
                if indicator is None:
                    CountryIndicator.objects.create(
                        country=country,
                        year=year,
                        safety_index=item["safety_index"],
                        cost_index=item["cost_index"],
                        tourism_index=item["tourism_index"],
                        climate_index=Decimal("50.00"),
                        medical_index=Decimal("50.00"),
                        visa_index=Decimal("50.00"),
                        overall_score=item["happiness_index"],
                        data_source=import_source,
                    )
                    created_indicators += 1
                else:
                    indicator.safety_index = item["safety_index"]
                    indicator.cost_index = item["cost_index"]
                    indicator.tourism_index = item["tourism_index"]
                    indicator.overall_score = item["happiness_index"]
                    indicator.data_source = import_source
                    indicator.save(
                        update_fields=[
                            "safety_index",
                            "cost_index",
                            "tourism_index",
                            "overall_score",
                            "data_source",
                            "updated_at",
                        ]
                    )
                    updated_indicators += 1

        record_operation_log(
            request,
            OperationLog.OPERATION_EXCEL_IMPORT,
            operation_object=excel_file.name,
            detail=(
                f"导入成功：年份 {year}，共 {len(rows)} 行，"
                f"新增国家 {created_countries} 个，更新国家 {updated_countries} 个，"
                f"新增指标 {created_indicators} 条，更新指标 {updated_indicators} 条"
            ),
        )
        return Response(
            {
                "code": 200,
                "message": "国家指标 Excel 导入成功",
                "data": {
                    "year": year,
                    "total_rows": len(rows),
                    "created_countries": created_countries,
                    "updated_countries": updated_countries,
                    "created_indicators": created_indicators,
                    "updated_indicators": updated_indicators,
                },
            },
            status=status.HTTP_200_OK,
        )


class LatestCountryMapDataAPIView(APIView):
    """首页世界地图所需的国家指标数据接口。"""

    permission_classes = [AllowAny]

    @extend_schema(
        tags=["国家数据"],
        summary="获取首页地图国家指标数据",
        description="返回最新年份的国家指标和推荐指数，用于前端世界地图可视化着色。",
        auth=[],
        responses={200: OpenApiTypes.OBJECT},
    )
    def get(self, request):
        """返回最新年份的国家指标数据，供前端地图可视化使用。"""
        latest_year = CountryIndicator.objects.order_by("-year").values_list(
            "year", flat=True
        ).first()
        if latest_year is None:
            return Response(
                {
                    "code": 200,
                    "message": "暂无国家指标数据",
                    "data": {
                        "year": None,
                        "results": [],
                    },
                },
                status=status.HTTP_200_OK,
            )

        indicators = CountryIndicator.objects.select_related("country").filter(
            year=latest_year,
            country__is_active=True,
        ).order_by("-tourism_index", "-safety_index", "country__name_en")

        results = []
        for indicator in indicators:
            results.append(build_indicator_payload(indicator))

        return Response(
            {
                "code": 200,
                "message": "获取地图数据成功",
                "data": {
                    "year": latest_year,
                    "results": results,
                },
            },
            status=status.HTTP_200_OK,
        )


class CountryContinentStatsAPIView(APIView):
    """按洲别统计国家指标数据接口。"""

    permission_classes = [IsAdminRole]

    @extend_schema(
        tags=["国家数据"],
        summary="按洲别统计国家指标",
        description="返回最新年份下各洲国家数量以及平均推荐指数、安全指数、消费指数等统计数据。",
        responses={200: OpenApiTypes.OBJECT},
    )
    def get(self, request):
        """返回最新年份下各洲国家数量与平均指标。"""
        latest_year = CountryIndicator.objects.order_by("-year").values_list(
            "year", flat=True
        ).first()
        if latest_year is None:
            return Response(
                {
                    "code": 200,
                    "message": "暂无国家指标数据",
                    "data": {"year": None, "results": []},
                },
                status=status.HTTP_200_OK,
            )

        indicators = CountryIndicator.objects.select_related("country").filter(
            year=latest_year,
            country__is_active=True,
        )
        stats_map = {}
        for indicator in indicators:
            payload = build_indicator_payload(indicator)
            continent = payload["continent"] or "未分类"
            if continent not in stats_map:
                stats_map[continent] = {
                    "continent": continent,
                    "country_count": 0,
                    "recommendation_sum": 0,
                    "tourism_sum": 0,
                    "safety_sum": 0,
                    "cost_sum": 0,
                    "happiness_sum": 0,
                }

            item = stats_map[continent]
            item["country_count"] += 1
            item["recommendation_sum"] += payload["recommendation_index"]
            item["tourism_sum"] += payload["tourism_index"]
            item["safety_sum"] += payload["safety_index"]
            item["cost_sum"] += payload["cost_index"]
            item["happiness_sum"] += payload["happiness_index"]

        results = []
        for item in stats_map.values():
            count = item["country_count"] or 1
            results.append(
                {
                    "continent": item["continent"],
                    "country_count": item["country_count"],
                    "avg_recommendation_index": round(item["recommendation_sum"] / count, 2),
                    "avg_tourism_index": round(item["tourism_sum"] / count, 2),
                    "avg_safety_index": round(item["safety_sum"] / count, 2),
                    "avg_cost_index": round(item["cost_sum"] / count, 2),
                    "avg_happiness_index": round(item["happiness_sum"] / count, 2),
                }
            )

        results.sort(key=lambda item: item["avg_recommendation_index"], reverse=True)
        return Response(
            {
                "code": 200,
                "message": "获取洲别统计成功",
                "data": {"year": latest_year, "results": results},
            },
            status=status.HTTP_200_OK,
        )


class CountryInsightDetailAPIView(APIView):
    """国家洞察详情接口。"""

    permission_classes = [IsAdminRole]

    @extend_schema(
        tags=["国家数据"],
        summary="获取国家洞察详情",
        description="返回指定国家基础信息、最新指标和历史年度指标，用于国家详情页可视化。",
        responses={200: OpenApiTypes.OBJECT, 404: OpenApiResponse(description="国家不存在或暂无指标数据")},
    )
    def get(self, request, pk):
        """返回指定国家的基础资料、最新指标和历史指标。"""
        country = get_object_or_404(Country.objects.filter(is_active=True), pk=pk)
        latest_indicator = (
            CountryIndicator.objects.select_related("country")
            .filter(country=country)
            .order_by("-year")
            .first()
        )
        if latest_indicator is None:
            return Response(
                {
                    "code": 404,
                    "message": "该国家暂无指标数据",
                    "data": None,
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        history = [
            build_indicator_payload(indicator)
            for indicator in CountryIndicator.objects.select_related("country")
            .filter(country=country)
            .order_by("year")
        ]
        return Response(
            {
                "code": 200,
                "message": "获取国家洞察详情成功",
                "data": {
                    "country": {
                        "id": country.id,
                        "name_zh": country.name_zh,
                        "name_en": country.name_en,
                        "code": country.code,
                        "continent": country.continent,
                        "capital": country.capital,
                        "language": country.language,
                        "currency": country.currency,
                    },
                    "latest_indicator": build_indicator_payload(latest_indicator),
                    "history": history,
                },
            },
            status=status.HTTP_200_OK,
        )


@extend_schema_view(
    list=extend_schema(tags=["国家数据"], summary="管理端查询国家数据"),
    retrieve=extend_schema(tags=["国家数据"], summary="管理端查看单个国家"),
    create=extend_schema(tags=["国家数据"], summary="添加国家数据", request=CountrySerializer, responses=CountrySerializer),
    update=extend_schema(tags=["国家数据"], summary="全量修改国家数据", request=CountrySerializer, responses=CountrySerializer),
    partial_update=extend_schema(tags=["国家数据"], summary="局部修改国家数据", request=CountrySerializer, responses=CountrySerializer),
    destroy=extend_schema(
        tags=["国家数据"],
        summary="删除国家数据",
        responses={204: OpenApiResponse(description="删除成功")},
    ),
)
class CountryAdminViewSet(ModelViewSet):
    """管理端国家数据增删改查接口。"""

    serializer_class = CountrySerializer
    permission_classes = [IsAdminRole]
    queryset = Country.objects.all().order_by("name_zh")

    def perform_create(self, serializer):
        """创建国家时记录操作日志。"""
        country = serializer.save()
        record_operation_log(
            self.request,
            OperationLog.OPERATION_COUNTRY_CREATE,
            operation_object=country.name_zh,
            detail=f"添加国家数据：{country.name_zh}",
        )

    def perform_update(self, serializer):
        """修改国家时记录操作日志。"""
        country = serializer.save()
        record_operation_log(
            self.request,
            OperationLog.OPERATION_COUNTRY_UPDATE,
            operation_object=country.name_zh,
            detail=f"修改国家数据：{country.name_zh}",
        )

    def perform_destroy(self, instance):
        """删除国家时记录操作日志。"""
        country_name = instance.name_zh
        instance.delete()
        record_operation_log(
            self.request,
            OperationLog.OPERATION_COUNTRY_DELETE,
            operation_object=country_name,
            detail=f"删除国家数据：{country_name}",
        )
