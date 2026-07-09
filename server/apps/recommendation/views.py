"""推荐模块视图。"""

from io import BytesIO
from urllib.parse import quote

from django.http import HttpResponse
from django.utils import timezone
from drf_spectacular.utils import (
    OpenApiExample,
    OpenApiParameter,
    OpenApiResponse,
    OpenApiTypes,
    extend_schema,
    inline_serializer,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import serializers, status
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.countries.models import CountryIndicator
from apps.recommendation.ai_explanation import generate_ai_recommendation_explanation
from apps.recommendation.models import RecommendationRecord, UserPreference
from apps.recommendation.recommendation import TravelRecommendationEngine
from apps.recommendation.serializers import RecommendationRequestSerializer
from apps.recommendation.tourism_suitability import TourismSuitabilityCalculator
from apps.system.models import OperationLog, record_operation_log
from common.permissions import IsAuthenticatedRole


RECOMMENDATION_EXPORT_HEADERS = [
    "排名",
    "国家名称",
    "英文名称",
    "所属洲别",
    "综合得分",
    "旅游适宜指数",
    "安全指数",
    "幸福指数",
    "消费指数",
    "预计消费",
    "安全需求",
    "安全匹配",
    "推荐说明",
    "数据年份",
    "导出时间",
]


def get_export_value(item, *keys, default=""):
    """按候选字段顺序读取推荐结果值。"""
    for key in keys:
        value = item.get(key)
        if value not in (None, ""):
            return value
    return default


def build_ai_country_payload(indicator):
    """将国家最新指标整理为 AI 推荐说明所需上下文。"""
    raw_visa_score = (
        float(indicator.visa_index)
        if float(indicator.visa_index) != 50.0
        else float(indicator.tourism_index)
    )
    country_payload = {
        "country_id": indicator.country.id,
        "country_name": indicator.country.name_zh,
        "country_name_en": indicator.country.name_en,
        "continent": indicator.country.continent,
        "raw_visa_score": raw_visa_score,
        "visa_index": float(indicator.visa_index),
        "tourism_index": float(indicator.tourism_index),
        "safety_index": float(indicator.safety_index),
        "cost_index": float(indicator.cost_index),
        "ppp_index": float(indicator.cost_index),
        "happiness_index": float(indicator.overall_score),
    }
    tourism_detail = TourismSuitabilityCalculator.build_detail(country_payload)
    country_payload["tourism_index"] = tourism_detail["tourism_index"]
    country_payload["tourism_detail"] = tourism_detail
    country_payload["recommendation_index"] = (
        TravelRecommendationEngine.calculate_default_recommendation_index(country_payload)
    )
    country_payload["year"] = indicator.year
    return country_payload


class RecommendationAPIView(APIView):
    """旅游国家推荐接口。"""

    # 推荐接口允许匿名访问
    permission_classes = [IsAuthenticatedRole]

    @extend_schema(
        tags=["推荐算法"],
        summary="生成安全旅游国家推荐",
        description=(
            "根据用户预算等级、偏好洲别和安全需求，使用固定权重推荐算法返回前 10 个推荐国家。"
            "已登录用户每次调用会保存 1 条偏好记录和多条推荐结果记录，用于后台推荐次数趋势统计。"
        ),
        request=RecommendationRequestSerializer,
        responses={
            200: OpenApiResponse(description="推荐成功，返回年份、推荐数量和推荐国家列表"),
            400: OpenApiResponse(description="暂无国家指标数据或请求参数不合法"),
        },
        examples=[
            OpenApiExample(
                "推荐请求示例",
                value={
                    "budget_level": "medium",
                    "preferred_continent": "亚洲",
                    "safety_requirement": "high",
                },
                request_only=True,
            )
        ],
    )
    def post(self, request):
        """接收预算和偏好条件，返回排序后的推荐国家列表。"""
        serializer = RecommendationRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated_data = serializer.validated_data

        # 获取最新年份的国家指标数据作为推荐基础
        latest_indicator_year = CountryIndicator.objects.order_by("-year").values_list(
            "year", flat=True
        ).first()
        if latest_indicator_year is None:
            return Response(
                {
                    "code": 400,
                    "message": "当前暂无国家指标数据，无法生成推荐结果",
                    "data": [],
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        indicators = CountryIndicator.objects.select_related("country").filter(
            year=latest_indicator_year,
            country__is_active=True,
        )
        if not indicators.exists():
            return Response(
                {
                    "code": 404,
                    "message": "未查询到符合条件的国家数据",
                    "data": [],
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        # 将数据库记录转换为独立推荐算法模块需要的输入格式
        countries_data = []
        for indicator in indicators:
            raw_visa_score = (
                float(indicator.visa_index)
                if float(indicator.visa_index) != 50.0
                else float(indicator.tourism_index)
            )
            country_payload = {
                "country_id": indicator.country.id,
                "country_name": indicator.country.name_zh,
                "country_name_en": indicator.country.name_en,
                "continent": indicator.country.continent,
                "raw_visa_score": raw_visa_score,
                "visa_index": float(indicator.visa_index),
                "tourism_index": float(indicator.tourism_index),
                "safety_index": float(indicator.safety_index),
                "cost_index": float(indicator.cost_index),
                "ppp_index": float(indicator.cost_index),
                "happiness_index": float(indicator.overall_score),
            }
            tourism_detail = TourismSuitabilityCalculator.build_detail(country_payload)
            country_payload["tourism_index"] = tourism_detail["tourism_index"]
            country_payload["tourism_detail"] = tourism_detail
            countries_data.append(
                country_payload
            )

        # 调用独立推荐算法模块完成综合评分与排序
        engine = TravelRecommendationEngine(
            countries=countries_data,
            user_preference=validated_data,
        )
        top_recommendations = engine.recommend(top_n=10)

        # 补充年份信息，方便前端展示和论文说明
        for item in top_recommendations:
            item["year"] = latest_indicator_year

        # 如果用户已登录，则保存本次偏好与推荐记录
        if request.user.is_authenticated:
            safety_requirement = validated_data.get("safety_requirement", "high")
            safety_requirement_text = {
                "normal": "一般安全需求",
                "high": "较高安全需求",
                "strict": "高安全需求",
            }.get(safety_requirement, "较高安全需求")
            preference = UserPreference.objects.create(
                user=request.user,
                preferred_continent=validated_data.get("preferred_continent", ""),
                budget_level=validated_data["budget_level"],
                # 以下字段仅为兼容当前数据表结构保留
                safety_weight=30,
                cost_weight=15,
                climate_weight=15,
                medical_weight=40,
                visa_weight=0,
                remark=(
                    "该记录由固定权重旅游推荐算法生成：旅游适宜性40%，安全30%，"
                    f"幸福15%，消费15%；安全需求为{safety_requirement_text}。"
                ),
            )
            for index, item in enumerate(top_recommendations, start=1):
                RecommendationRecord.objects.create(
                    user=request.user,
                    preference=preference,
                    country_id=item["country_id"],
                    score=item["score"],
                    reason=item["reason"],
                    rank=index,
                )

        return Response(
            {
                "code": 200,
                "message": "推荐成功",
                "data": {
                    "year": latest_indicator_year,
                    "count": len(top_recommendations),
                    "results": top_recommendations,
                },
            },
            status=status.HTTP_200_OK,
        )


class RecommendationExportExcelAPIView(APIView):
    """推荐结果 Excel 导出接口。"""

    permission_classes = [IsAuthenticatedRole]

    @extend_schema(
        tags=["推荐算法"],
        summary="导出推荐结果 Excel",
        description="将当前推荐结果列表导出为 .xlsx 文件，并记录 Excel 导出操作日志。",
        request=inline_serializer(
            name="RecommendationExportExcelRequest",
            fields={
                "year": serializers.CharField(required=False, help_text="推荐结果对应的数据年份"),
                "results": serializers.ListField(
                    child=serializers.DictField(),
                    help_text="前端当前展示的推荐结果列表",
                ),
            },
        ),
        responses={
            (200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"): OpenApiTypes.BINARY,
            400: OpenApiResponse(description="暂无可导出的推荐结果或结果格式错误"),
        },
        examples=[
            OpenApiExample(
                "导出请求示例",
                value={
                    "year": 2026,
                    "results": [
                        {
                            "rank": 1,
                            "country_name": "日本",
                            "country_name_en": "Japan",
                            "continent": "亚洲",
                            "score": 88.5,
                            "tourism_index": 91,
                            "safety_index": 88,
                            "happiness_index": 74,
                            "cost_index": 62,
                            "reason": "安全指数和旅游适宜指数较高",
                        }
                    ],
                },
                request_only=True,
            )
        ],
    )
    def post(self, request):
        """根据当前推荐结果生成 Excel 文件。"""
        results = request.data.get("results") or []
        year = request.data.get("year") or ""
        if not isinstance(results, list) or not results:
            record_operation_log(
                request,
                OperationLog.OPERATION_EXCEL_EXPORT,
                operation_object="推荐结果 Excel",
                operation_result=OperationLog.RESULT_FAILED,
                detail="导出失败：暂无可导出的推荐结果",
            )
            return Response(
                {
                    "code": 400,
                    "message": "暂无可导出的推荐结果",
                    "data": None,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        if any(not isinstance(item, dict) for item in results):
            record_operation_log(
                request,
                OperationLog.OPERATION_EXCEL_EXPORT,
                operation_object="推荐结果 Excel",
                operation_result=OperationLog.RESULT_FAILED,
                detail="导出失败：推荐结果格式不正确",
            )
            return Response(
                {
                    "code": 400,
                    "message": "推荐结果格式不正确",
                    "data": None,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        export_time = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "推荐结果"
        worksheet.append(RECOMMENDATION_EXPORT_HEADERS)

        header_fill = PatternFill("solid", fgColor="2F6B5A")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        wrap_alignment = Alignment(vertical="top", wrap_text=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        for index, item in enumerate(results, start=1):
            safety_matched = get_export_value(item, "safety_matched", default=True)
            if isinstance(safety_matched, bool):
                safety_matched_text = "满足" if safety_matched else "未完全满足"
            else:
                safety_matched_text = str(safety_matched)

            worksheet.append(
                [
                    get_export_value(item, "rank", default=index),
                    get_export_value(item, "country_name", "country_name_zh"),
                    get_export_value(item, "country_name_en"),
                    get_export_value(item, "continent"),
                    get_export_value(item, "score"),
                    get_export_value(item, "tourism_index"),
                    get_export_value(item, "safety_index"),
                    get_export_value(item, "happiness_index", "overall_score"),
                    get_export_value(item, "ppp_index", "cost_index"),
                    get_export_value(item, "estimated_cost"),
                    get_export_value(item, "safety_requirement"),
                    safety_matched_text,
                    get_export_value(item, "reason"),
                    year,
                    export_time,
                ]
            )

        column_widths = {
            "A": 10,
            "B": 18,
            "C": 18,
            "D": 14,
            "E": 12,
            "F": 16,
            "G": 12,
            "H": 12,
            "I": 12,
            "J": 12,
            "K": 16,
            "L": 14,
            "M": 36,
            "N": 12,
            "O": 20,
        }
        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap_alignment
            for cell in row[:12]:
                cell.alignment = center_alignment

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        filename = f"旅游推荐结果_{year or timezone.localdate().isoformat()}.xlsx"
        record_operation_log(
            request,
            OperationLog.OPERATION_EXCEL_EXPORT,
            operation_object=filename,
            detail=f"导出推荐结果：共 {len(results)} 条",
        )
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return response


class RecommendationExplanationAPIView(APIView):
    """单个国家 AI 推荐说明接口。"""

    permission_classes = [IsAuthenticatedRole]

    @extend_schema(
        tags=["推荐算法"],
        summary="生成单个国家 AI 推荐说明",
        description=(
            "根据指定国家最新指标生成推荐理由、安全提醒、消费水平说明和适合人群。"
            "后端可接入 DeepSeek/OpenAI；没有 API Key 时使用本地模板。"
        ),
        parameters=[
            OpenApiParameter("country_id", int, OpenApiParameter.PATH, description="国家 ID"),
        ],
        responses={
            200: OpenApiResponse(description="推荐说明生成成功"),
            404: OpenApiResponse(description="国家不存在或暂无指标数据"),
        },
    )
    def get(self, request, country_id):
        """根据国家最新指标生成推荐说明。"""
        indicator = (
            CountryIndicator.objects.select_related("country")
            .filter(country_id=country_id, country__is_active=True)
            .order_by("-year")
            .first()
        )
        if indicator is None:
            return Response(
                {
                    "code": 404,
                    "message": "该国家暂无指标数据，无法生成推荐说明",
                    "data": None,
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        country_payload = build_ai_country_payload(indicator)
        explanation_result = generate_ai_recommendation_explanation(country_payload)

        return Response(
            {
                "code": 200,
                "message": "推荐说明生成成功",
                "data": {
                    "country": {
                        "id": indicator.country.id,
                        "name_zh": indicator.country.name_zh,
                        "name_en": indicator.country.name_en,
                        "continent": indicator.country.continent,
                    },
                    "indicator": country_payload,
                    **explanation_result,
                },
            },
            status=status.HTTP_200_OK,
        )
